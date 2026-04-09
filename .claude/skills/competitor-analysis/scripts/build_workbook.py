#!/usr/bin/env python3
"""
Competitor Analysis Workbook Builder

Builds an 8-tab Excel workbook from Parcl Labs Portfolio Hunter CSV exports.

Usage:
    python3 build_workbook.py \
        --msa-activity amh-msa-activity.csv \
        --properties amh-properties.csv \
        --rentals amh-rentals.csv \
        --for-sale amh-for-sale.csv \
        --transactions amh-transactions.csv \
        --rpi amh-rpi.csv \
        --output amh-competitor-analysis.xlsx \
        --investor "AMH"
"""

import argparse
import csv
import statistics
import sys
from collections import defaultdict
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Styles ────────────────────────────────────────────────────────────────────

NAVY = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
LIGHT_BLUE = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

WHITE_BOLD_14 = Font(name="Arial", size=14, bold=True, color="FFFFFF")
WHITE_BOLD_10 = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BOLD_10 = Font(name="Arial", size=10, bold=True)
NORMAL_10 = Font(name="Arial", size=10)
GREEN_10 = Font(name="Arial", size=10, color="006100")
RED_10 = Font(name="Arial", size=10, color="9C0006")
GRAY_10 = Font(name="Arial", size=10, color="666666")

THIN_BORDER = Border(
    bottom=Side(style="thin", color="D9D9D9"),
)


# ── Helpers ───────────────────────────────────────────────────────────────────

def clean_msa_id(val):
    """Normalize MSA parcl_id: strip .0 suffix, return as string."""
    if val is None:
        return ""
    s = str(val).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def safe_float(val, default=0.0):
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def safe_int(val, default=0):
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def percentile(data, p):
    if not data:
        return 0
    s = sorted(data)
    k = (len(s) - 1) * p / 100
    f = int(k)
    c = min(f + 1, len(s) - 1)
    return s[f] + (k - f) * (s[c] - s[f])


def write_title(ws, text, max_col, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    ws.cell(row, 1, text).font = WHITE_BOLD_14
    for c in range(1, max_col + 1):
        ws.cell(row, c).fill = NAVY


def write_headers(ws, headers, row=2):
    for i, h in enumerate(headers):
        cell = ws.cell(row, i + 1, h)
        cell.font = WHITE_BOLD_10
        cell.fill = NAVY


def activity_font(activity_type):
    if "Buyer" in str(activity_type):
        return GREEN_10
    if "Seller" in str(activity_type):
        return RED_10
    return NORMAL_10


def set_col_widths(ws, widths):
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w


# ── Known MSA names (fallback lookup) ─────────────────────────────────────────

KNOWN_MSAS = {
    "2887280": "Atlanta-Sandy Springs-Alpharetta, GA",
    "2899841": "Charlotte-Concord-Gastonia, NC-SC",
    "2899734": "Dallas-Fort Worth-Arlington, TX",
    "2900174": "Nashville-Davidson-Murfreesboro, TN",
    "2899989": "Jacksonville, FL",
    "2900245": "Phoenix-Mesa-Chandler, AZ",
    "2899979": "Indianapolis-Carmel-Anderson, IN",
    "2900417": "Tampa-St Petersburg-Clearwater, FL",
    "2899967": "Houston-The Woodlands-Sugar Land, TX",
    "2900049": "Las Vegas-Henderson-Paradise, NV",
    "2900282": "Raleigh-Cary, NC",
    "2899621": "Boise City, ID",
    "2900266": "Portland-Vancouver-Hillsboro, OR-WA",
    "2900306": "Salt Lake City, UT",
    "2899633": "Denver-Aurora-Lakewood, CO",
    "2900310": "San Antonio-New Braunfels, TX",
    "2900153": "Orlando-Kissimmee-Sanford, FL",
    "2899615": "Austin-Round Rock-Georgetown, TX",
    "2900372": "Sacramento-Roseville-Folsom, CA",
}


def resolve_msa_names(properties_rows):
    """Build MSA ID -> name mapping from property city/state data."""
    msa_to_state = defaultdict(lambda: defaultdict(int))
    for row in properties_rows:
        mid = clean_msa_id(row.get("msa_parcl_id"))
        state = row.get("enhanced_address_state", "")
        city = row.get("enhanced_address_city", "")
        if mid and state:
            msa_to_state[mid][f"{city}, {state}"] += 1

    names = {}
    for mid, cities in msa_to_state.items():
        if mid in KNOWN_MSAS:
            names[mid] = KNOWN_MSAS[mid]
        else:
            top_city = max(cities, key=cities.get) if cities else f"MSA {mid}"
            names[mid] = top_city
    # Fill in any known MSAs not in property data
    for mid, name in KNOWN_MSAS.items():
        if mid not in names:
            names[mid] = name
    return names


# ── Main ──────────────────────────────────────────────────────────────────────

def build_workbook(args):
    investor_name = args.investor.upper()
    today = date.today().strftime("%B %d, %Y")

    # Load CSVs
    msa_rows = list(csv.DictReader(open(args.msa_activity)))
    props_rows = list(csv.DictReader(open(args.properties)))
    rentals_rows = list(csv.DictReader(open(args.rentals)))
    forsale_rows = list(csv.DictReader(open(args.for_sale)))
    txn_rows = list(csv.DictReader(open(args.transactions)))
    rpi_rows = list(csv.DictReader(open(args.rpi)))

    # Normalize MSA IDs everywhere
    for r in msa_rows:
        r["msa_parcl_id"] = clean_msa_id(r.get("msa_parcl_id"))
    for r in props_rows:
        r["msa_parcl_id"] = clean_msa_id(r.get("msa_parcl_id"))
    for r in rentals_rows:
        r["msa_parcl_id"] = clean_msa_id(r.get("msa_parcl_id"))
    for r in forsale_rows:
        r["msa_parcl_id"] = clean_msa_id(r.get("msa_parcl_id"))
    for r in txn_rows:
        r["msa_parcl_id"] = clean_msa_id(r.get("msa_parcl_id"))
    for r in rpi_rows:
        r["parcl_id"] = clean_msa_id(r.get("parcl_id"))

    # Sort MSAs by property count
    msa_rows.sort(key=lambda x: -safe_int(x.get("num_properties")))
    total_props = sum(safe_int(r["num_properties"]) for r in msa_rows)
    top10_ids = set(r["msa_parcl_id"] for r in msa_rows[:10])

    # Resolve names
    msa_names = resolve_msa_names(props_rows)

    def mname(mid):
        return msa_names.get(str(mid), f"MSA {mid}")

    # Index data by MSA
    props_by_msa = defaultdict(list)
    for r in props_rows:
        props_by_msa[r["msa_parcl_id"]].append(r)

    active_rentals = [r for r in rentals_rows if r.get("rental_status") == "Active Rental"]
    rentals_by_msa = defaultdict(list)
    for r in active_rentals:
        rentals_by_msa[r["msa_parcl_id"]].append(r)

    all_rentals_by_msa = defaultdict(list)
    for r in rentals_rows:
        all_rentals_by_msa[r["msa_parcl_id"]].append(r)

    fs_by_msa = defaultdict(list)
    for r in forsale_rows:
        fs_by_msa[r["msa_parcl_id"]].append(r)

    # Property sqft lookup for rent/sqft calculations
    prop_sqft = {}
    for r in props_rows:
        pid = r.get("parcl_property_id", "")
        sqft = safe_float(r.get("sq_ft"), 0)
        if pid and sqft > 0:
            prop_sqft[pid] = sqft

    wb = Workbook()

    # ── TAB 1: Portfolio Overview ─────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Portfolio Overview"
    write_title(ws1, f"{investor_name} PORTFOLIO OVERVIEW", 5)

    ws1.cell(2, 1, f"Data as of {today} | Source: Parcl Labs").font = GRAY_10

    ws1.merge_cells("A4:C4")
    ws1.cell(4, 1, "NATIONAL SUMMARY").font = BOLD_10
    for c in range(1, 4):
        ws1.cell(4, c).fill = LIGHT_BLUE

    portfolio_value = sum(safe_float(r.get("portfolio_value")) for r in msa_rows)
    avg_hold = None
    national_summary = [
        ("Total Properties", f"{total_props:,}"),
        ("Estimated Portfolio Value", f"${portfolio_value:,.0f}"),
        ("Number of MSAs", len(msa_rows)),
        ("Active Rental Listings", len(active_rentals)),
        ("For-Sale Listings", len(forsale_rows)),
    ]
    for i, (lbl, val) in enumerate(national_summary):
        ws1.cell(5 + i, 1, lbl).font = BOLD_10
        ws1.cell(5 + i, 2, val).font = NORMAL_10

    # Market breakdown
    ws1.cell(12, 1, "").font = NORMAL_10
    write_headers(ws1, ["Market", "Properties", "% of Portfolio", "Est. Value ($M)", "12-Mo Activity"], row=13)
    for ri, r in enumerate(msa_rows):
        row = 14 + ri
        ws1.cell(row, 1, mname(r["msa_parcl_id"])).font = NORMAL_10
        ws1.cell(row, 2, safe_int(r["num_properties"])).font = NORMAL_10
        ws1.cell(row, 3, safe_int(r["num_properties"]) / total_props if total_props else 0).font = NORMAL_10
        ws1.cell(row, 3).number_format = "0.0%"
        pv = safe_float(r.get("portfolio_value")) / 1_000_000
        ws1.cell(row, 4, round(pv, 1)).font = NORMAL_10
        ws1.cell(row, 4).number_format = "$#,##0.0"
        ws1.cell(row, 5, r.get("activity_type", "")).font = activity_font(r.get("activity_type", ""))

        # Bold top 10
        if r["msa_parcl_id"] in top10_ids:
            ws1.cell(row, 1).font = BOLD_10

    # National totals row
    totals_row = 14 + len(msa_rows)
    ws1.cell(totals_row, 1, "NATIONAL TOTAL").font = BOLD_10
    ws1.cell(totals_row, 1).fill = LIGHT_BLUE
    ws1.cell(totals_row, 2, total_props).font = BOLD_10
    ws1.cell(totals_row, 2).fill = LIGHT_BLUE
    ws1.cell(totals_row, 3, 1.0).font = BOLD_10
    ws1.cell(totals_row, 3).number_format = "0.0%"
    ws1.cell(totals_row, 3).fill = LIGHT_BLUE
    ws1.cell(totals_row, 4, round(portfolio_value / 1_000_000, 1)).font = BOLD_10
    ws1.cell(totals_row, 4).number_format = "$#,##0.0"
    ws1.cell(totals_row, 4).fill = LIGHT_BLUE
    for c in range(5, 6):
        ws1.cell(totals_row, c).fill = LIGHT_BLUE

    set_col_widths(ws1, [45, 14, 14, 16, 16])
    ws1.freeze_panes = "A14"

    # ── TAB 2: Rental Price Index ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Rental Price Index")

    rpi_by_market = defaultdict(dict)
    for r in rpi_rows:
        mid = r["parcl_id"]
        if mid in top10_ids:
            mo = r["date"][:7]
            rpi_by_market[mid][mo] = safe_float(r.get("value"))

    rpi_markets = [mid for mid in [mr["msa_parcl_id"] for mr in msa_rows[:10]] if mid in rpi_by_market]
    months = sorted(set(mo for mkt in rpi_by_market.values() for mo in mkt))

    ncols = len(rpi_markets) + 1
    write_title(ws2, f"{investor_name} TOP MARKETS: RENTAL PRICE INDEX ($/SQFT/MO)", ncols)

    ws2.cell(2, 1, "Month").font = WHITE_BOLD_10
    ws2.cell(2, 1).fill = NAVY
    for i, mid in enumerate(rpi_markets):
        c = ws2.cell(2, i + 2, mname(mid)[:25])
        c.font = WHITE_BOLD_10
        c.fill = NAVY

    for ri, mo in enumerate(months):
        ws2.cell(3 + ri, 1, mo).font = NORMAL_10
        for ci, mid in enumerate(rpi_markets):
            val = rpi_by_market[mid].get(mo)
            if val is not None:
                cell = ws2.cell(3 + ri, ci + 2, round(val, 2))
                cell.font = NORMAL_10
                cell.number_format = "$#,##0.00"

    ws2.column_dimensions["A"].width = 12
    for i in range(len(rpi_markets)):
        ws2.column_dimensions[get_column_letter(i + 2)].width = 28
    ws2.freeze_panes = "B3"

    # ── TAB 3: Buy Box ────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Buy Box")
    bb_headers = [
        "Market", "Units", "P10 SqFt", "Med SqFt", "P90 SqFt",
        "Med Beds", "Med Baths", "P10 Year Built", "Med Year Built",
        "P90 Year Built", "Med Acq Price",
    ]
    write_title(ws3, f"{investor_name} BUY BOX BY MARKET", len(bb_headers))
    write_headers(ws3, bb_headers)

    row = 3
    all_sqfts, all_beds, all_baths, all_years, all_acq = [], [], [], [], []
    for mr in msa_rows[:10]:
        mid = mr["msa_parcl_id"]
        plist = props_by_msa.get(mid, [])
        if not plist:
            continue

        sqfts = [safe_float(p["sq_ft"]) for p in plist if safe_float(p.get("sq_ft")) > 0]
        beds = [safe_float(p["bedrooms"]) for p in plist if safe_float(p.get("bedrooms")) > 0]
        baths = [safe_float(p["bathrooms"]) for p in plist if safe_float(p.get("bathrooms")) > 0]
        years = [safe_int(p["year_built"]) for p in plist if safe_int(p.get("year_built")) > 1800]
        acq = [safe_float(p["acquisition_price"]) for p in plist if safe_float(p.get("acquisition_price")) > 0]

        all_sqfts.extend(sqfts)
        all_beds.extend(beds)
        all_baths.extend(baths)
        all_years.extend(years)
        all_acq.extend(acq)

        vals = [
            mname(mid), len(plist),
            round(percentile(sqfts, 10)), round(statistics.median(sqfts)) if sqfts else 0, round(percentile(sqfts, 90)),
            round(statistics.median(beds)) if beds else 0, round(statistics.median(baths), 1) if baths else 0,
            round(percentile(years, 10)), round(statistics.median(years)) if years else 0, round(percentile(years, 90)),
            round(statistics.median(acq)) if acq else 0,
        ]
        fmts = [None, "#,##0", "#,##0", "#,##0", "#,##0", "0", "0.0", "0", "0", "0", "$#,##0"]
        for ci, (v, fmt) in enumerate(zip(vals, fmts)):
            cell = ws3.cell(row, ci + 1, v)
            cell.font = BOLD_10 if ci == 0 else NORMAL_10
            if fmt:
                cell.number_format = fmt
        row += 1

    # National rollup
    nat_vals = [
        "NATIONAL ROLLUP", sum(len(props_by_msa[mid]) for mid in top10_ids),
        round(percentile(all_sqfts, 10)), round(statistics.median(all_sqfts)) if all_sqfts else 0, round(percentile(all_sqfts, 90)),
        round(statistics.median(all_beds)) if all_beds else 0, round(statistics.median(all_baths), 1) if all_baths else 0,
        round(percentile(all_years, 10)), round(statistics.median(all_years)) if all_years else 0, round(percentile(all_years, 90)),
        round(statistics.median(all_acq)) if all_acq else 0,
    ]
    for ci, (v, fmt) in enumerate(zip(nat_vals, fmts)):
        cell = ws3.cell(row, ci + 1, v)
        cell.font = BOLD_10
        cell.fill = LIGHT_BLUE
        if fmt:
            cell.number_format = fmt

    set_col_widths(ws3, [40, 10, 10, 10, 10, 10, 10, 12, 12, 12, 14])
    ws3.freeze_panes = "A3"

    # ── TAB 4: Acquisitions vs Sales ──────────────────────────────────────────
    ws4 = wb.create_sheet("Acquisitions vs Sales")
    txn_headers = ["Market", "Year", "Purchases", "Sales", "Net", "Avg Purchase $", "Avg Sale $"]
    write_title(ws4, f"{investor_name} ACQUISITIONS VS SALES", len(txn_headers))
    write_headers(ws4, txn_headers)

    txn_by_key = defaultdict(lambda: {"purchased": [], "sold": []})
    for t in txn_rows:
        mid = t["msa_parcl_id"]
        yr = t.get("event_date", "")[:4]
        tp = t.get("transaction_type", "")
        price = safe_float(t.get("transaction_price"))
        if mid in top10_ids and yr >= "2018" and tp in ("purchased", "sold"):
            txn_by_key[(mid, yr)][tp].append(price)

    row = 3
    for mr in msa_rows[:10]:
        mid = mr["msa_parcl_id"]
        years = sorted(set(yr for (m, yr) in txn_by_key if m == mid))
        for yr in years:
            d = txn_by_key[(mid, yr)]
            pc, sc = len(d["purchased"]), len(d["sold"])
            net = pc - sc
            avg_p = statistics.mean(d["purchased"]) if d["purchased"] else 0
            avg_s = statistics.mean(d["sold"]) if d["sold"] else 0

            ws4.cell(row, 1, mname(mid)[:35]).font = NORMAL_10
            ws4.cell(row, 2, int(yr)).font = NORMAL_10
            ws4.cell(row, 3, pc).font = NORMAL_10
            ws4.cell(row, 4, sc).font = NORMAL_10
            net_cell = ws4.cell(row, 5, net)
            net_cell.font = GREEN_10 if net > 0 else RED_10 if net < 0 else NORMAL_10
            ws4.cell(row, 6, round(avg_p)).font = NORMAL_10
            ws4.cell(row, 6).number_format = "$#,##0"
            ws4.cell(row, 7, round(avg_s)).font = NORMAL_10
            ws4.cell(row, 7).number_format = "$#,##0"
            row += 1

    set_col_widths(ws4, [38, 8, 12, 10, 8, 16, 16])
    ws4.freeze_panes = "A3"

    # ── TAB 5: Availability & Disposition ─────────────────────────────────────
    ws5 = wb.create_sheet("Availability")
    avail_headers = [
        "Market", "Total Props", "Active Rentals", "% For Rent",
        "Median Rent", "Avg Rental DOM", "For-Sale", "% For Sale",
        "Median List Price", "Avg Price Cuts", "Avg Unrealized P&L",
    ]
    write_title(ws5, f"{investor_name} AVAILABILITY & DISPOSITION", len(avail_headers))
    write_headers(ws5, avail_headers)

    row = 3
    for mr in msa_rows[:10]:
        mid = mr["msa_parcl_id"]
        n = safe_int(mr["num_properties"])
        ar = rentals_by_msa.get(mid, [])
        fs = fs_by_msa.get(mid, [])

        rents = [safe_float(r["last_rental_price"]) for r in ar if safe_float(r.get("last_rental_price")) > 0]
        rdoms = [safe_float(r["rental_days_on_market"]) for r in ar if r.get("rental_days_on_market")]
        fprices = [safe_float(r["latest_listing_price"]) for r in fs if safe_float(r.get("latest_listing_price")) > 0]
        fcuts = [safe_float(r["num_price_cuts"]) for r in fs if r.get("num_price_cuts")]
        fpnl = [safe_float(r["listing_unrealized_pnl"]) for r in fs if r.get("listing_unrealized_pnl")]

        vals = [
            mname(mid), n, len(ar), len(ar) / n if n else 0,
            statistics.median(rents) if rents else 0,
            round(statistics.mean(rdoms)) if rdoms else 0,
            len(fs), len(fs) / n if n else 0,
            statistics.median(fprices) if fprices else 0,
            round(statistics.mean(fcuts), 1) if fcuts else 0,
            round(statistics.mean(fpnl)) if fpnl else 0,
        ]
        fmts = [None, "#,##0", "#,##0", "0.0%", "$#,##0", "0", "#,##0", "0.0%", "$#,##0", "0.0", "$#,##0"]
        for ci, (v, fmt) in enumerate(zip(vals, fmts)):
            cell = ws5.cell(row, ci + 1, v)
            cell.font = BOLD_10 if ci == 0 else NORMAL_10
            if fmt:
                cell.number_format = fmt
        # Color P&L
        pnl_cell = ws5.cell(row, 11)
        if fpnl:
            avg_pnl = statistics.mean(fpnl)
            pnl_cell.font = GREEN_10 if avg_pnl > 0 else RED_10
        row += 1

    set_col_widths(ws5, [40, 12, 14, 10, 12, 12, 10, 10, 14, 10, 16])
    ws5.freeze_panes = "A3"

    # ── TAB 6: For-Sale Detail ────────────────────────────────────────────────
    ws6 = wb.create_sheet("For-Sale Detail")
    fs_headers = [
        "Market", "Address", "City", "State", "ZIP",
        "List Price", "DOM", "Price Cuts", "Unrealized P&L",
        "Agent", "Brokerage",
    ]
    write_title(ws6, f"{investor_name} FOR-SALE LISTINGS", len(fs_headers))
    write_headers(ws6, fs_headers)

    sorted_fs = sorted(forsale_rows, key=lambda x: -safe_float(x.get("latest_listing_price")))
    for ri, r in enumerate(sorted_fs):
        row = 3 + ri
        pnl = safe_float(r.get("listing_unrealized_pnl"))
        vals = [
            mname(r["msa_parcl_id"]),
            r.get("enhanced_address_address1", ""),
            r.get("enhanced_address_city", ""),
            r.get("enhanced_address_state", ""),
            r.get("enhanced_address_zip5", ""),
            safe_float(r.get("latest_listing_price")),
            safe_int(r.get("listing_days_on_market")),
            safe_int(r.get("num_price_cuts")),
            pnl,
            r.get("enhanced_agent_agent_name", ""),
            r.get("enhanced_agent_agent_business", ""),
        ]
        fmts = [None, None, None, None, None, "$#,##0", "0", "0", "$#,##0", None, None]
        for ci, (v, fmt) in enumerate(zip(vals, fmts)):
            cell = ws6.cell(row, ci + 1, v)
            cell.font = NORMAL_10
            if fmt:
                cell.number_format = fmt
        ws6.cell(row, 9).font = GREEN_10 if pnl > 0 else RED_10

    set_col_widths(ws6, [30, 35, 16, 6, 8, 14, 8, 10, 14, 25, 35])
    ws6.freeze_panes = "A3"
    ws6.auto_filter.ref = f"A2:K{2 + len(sorted_fs)}"

    # ── TAB 7: Investor vs Market Rents ───────────────────────────────────────
    ws7 = wb.create_sheet("Investor vs Market Rents")
    rent_headers = [
        "Market", "Active Listings", "Median Total Rent", "Median SqFt",
        "Investor $/SqFt/Mo", "Market $/SqFt/Mo (RPI)",
        "Premium/Discount %", "Avg Price Changes", "Avg DOM",
    ]
    write_title(ws7, f"{investor_name} VS MARKET RENTS", len(rent_headers))
    write_headers(ws7, rent_headers)

    # Latest RPI per market
    latest_rpi = {}
    for r in rpi_rows:
        mid = r["parcl_id"]
        latest_rpi[mid] = safe_float(r.get("value"))

    row = 3
    for mr in msa_rows[:10]:
        mid = mr["msa_parcl_id"]
        ar = rentals_by_msa.get(mid, [])
        if not ar:
            continue

        # Join with property sqft
        rents_with_sqft = []
        for r in ar:
            pid = r.get("parcl_property_id", "")
            rent = safe_float(r.get("last_rental_price"))
            sqft = prop_sqft.get(pid, 0)
            if rent > 0 and sqft > 0:
                rents_with_sqft.append((rent, sqft))

        all_rents = [safe_float(r["last_rental_price"]) for r in ar if safe_float(r.get("last_rental_price")) > 0]
        all_sqfts_r = [s for _, s in rents_with_sqft]
        per_sqft = [r / s for r, s in rents_with_sqft] if rents_with_sqft else []
        doms = [safe_float(r["rental_days_on_market"]) for r in ar if r.get("rental_days_on_market")]
        changes = [safe_float(r["rental_num_price_changes"]) for r in ar if r.get("rental_num_price_changes") is not None]

        med_rent = statistics.median(all_rents) if all_rents else 0
        med_sqft = statistics.median(all_sqfts_r) if all_sqfts_r else 0
        inv_psf = statistics.median(per_sqft) if per_sqft else 0
        mkt_psf = latest_rpi.get(mid, 0)
        premium = ((inv_psf / mkt_psf) - 1) * 100 if mkt_psf > 0 and inv_psf > 0 else 0

        vals = [
            mname(mid), len(ar), med_rent, round(med_sqft),
            round(inv_psf, 2), round(mkt_psf, 2),
            round(premium, 1),
            round(statistics.mean(changes), 1) if changes else 0,
            round(statistics.mean(doms)) if doms else 0,
        ]
        fmts = [None, "#,##0", "$#,##0", "#,##0", "$#,##0.00", "$#,##0.00", "+0.0%;-0.0%", "0.0", "0"]
        for ci, (v, fmt) in enumerate(zip(vals, fmts)):
            cell = ws7.cell(row, ci + 1, v)
            cell.font = BOLD_10 if ci == 0 else NORMAL_10
            if fmt:
                if fmt == "+0.0%;-0.0%":
                    cell.value = v / 100  # store as decimal for percentage format
                cell.number_format = fmt

        # Highlight premium/discount > 4%
        prem_cell = ws7.cell(row, 7)
        if abs(premium) > 4:
            prem_cell.fill = YELLOW_FILL
        row += 1

    set_col_widths(ws7, [40, 14, 14, 12, 16, 18, 16, 14, 10])
    ws7.freeze_panes = "A3"

    # ── TAB 8: Algorithmic Pricing ────────────────────────────────────────────
    ws8 = wb.create_sheet("Algorithmic Pricing")
    algo_headers = [
        "Market", "Active Listings", "Algo-Priced (3+)", "Algo %",
        "Avg Price Changes", "Med Price Changes",
        "Static (0 changes)", "Static %",
    ]
    write_title(ws8, f"{investor_name} ALGORITHMIC PRICING ANALYSIS", len(algo_headers))
    write_headers(ws8, algo_headers)

    row = 3
    for mr in msa_rows[:10]:
        mid = mr["msa_parcl_id"]
        ar = rentals_by_msa.get(mid, [])
        if not ar:
            continue

        changes = [safe_float(r.get("rental_num_price_changes", 0)) for r in ar]
        algo_3plus = sum(1 for c in changes if c >= 3)
        static = sum(1 for c in changes if c == 0)
        algo_pct = algo_3plus / len(ar) if ar else 0

        vals = [
            mname(mid), len(ar), algo_3plus, algo_pct,
            round(statistics.mean(changes), 1) if changes else 0,
            round(statistics.median(changes)) if changes else 0,
            static, static / len(ar) if ar else 0,
        ]
        fmts = [None, "#,##0", "#,##0", "0.0%", "0.0", "0", "#,##0", "0.0%"]
        for ci, (v, fmt) in enumerate(zip(vals, fmts)):
            cell = ws8.cell(row, ci + 1, v)
            cell.font = BOLD_10 if ci == 0 else NORMAL_10
            if fmt:
                cell.number_format = fmt

        # Color algo %
        pct_cell = ws8.cell(row, 4)
        if algo_pct > 0.9:
            pct_cell.fill = GREEN_FILL
        elif algo_pct > 0.8:
            pct_cell.fill = YELLOW_FILL
        row += 1

    # Frequency distribution (national)
    row += 2
    ws8.cell(row, 1, "PRICE CHANGE FREQUENCY DISTRIBUTION (NATIONAL)").font = BOLD_10
    ws8.cell(row, 1).fill = LIGHT_BLUE
    for c in range(2, 6):
        ws8.cell(row, c).fill = LIGHT_BLUE
    row += 1
    freq_headers = ["Band", "Count", "% of Active", "Cumulative %"]
    for i, h in enumerate(freq_headers):
        ws8.cell(row, i + 1, h).font = BOLD_10
    row += 1

    all_changes = [safe_float(r.get("rental_num_price_changes", 0)) for r in active_rentals]
    bands = [(0, 0, "0"), (1, 2, "1-2"), (3, 5, "3-5"), (6, 10, "6-10"), (11, 20, "11-20"), (21, 50, "21-50")]
    cum = 0
    for lo, hi, label in bands:
        cnt = sum(1 for c in all_changes if lo <= c <= hi)
        pct = cnt / len(all_changes) if all_changes else 0
        cum += pct
        ws8.cell(row, 1, label).font = NORMAL_10
        ws8.cell(row, 2, cnt).font = NORMAL_10
        ws8.cell(row, 3, pct).font = NORMAL_10
        ws8.cell(row, 3).number_format = "0.0%"
        ws8.cell(row, 4, cum).font = NORMAL_10
        ws8.cell(row, 4).number_format = "0.0%"
        row += 1

    set_col_widths(ws8, [40, 14, 14, 10, 16, 16, 14, 10])
    ws8.freeze_panes = "A3"

    # Save
    wb.save(args.output)
    print(f"Workbook saved: {args.output}")
    print(f"  Tab 1: Portfolio Overview ({len(msa_rows)} MSAs, {total_props:,} properties)")
    print(f"  Tab 2: Rental Price Index ({len(rpi_markets)} markets, {len(months)} months)")
    print(f"  Tab 3: Buy Box ({sum(1 for mr in msa_rows[:10] if mr['msa_parcl_id'] in props_by_msa)} markets)")
    print(f"  Tab 4: Acquisitions vs Sales")
    print(f"  Tab 5: Availability & Disposition")
    print(f"  Tab 6: For-Sale Detail ({len(forsale_rows)} listings)")
    print(f"  Tab 7: Investor vs Market Rents (with $/sqft join)")
    print(f"  Tab 8: Algorithmic Pricing (with frequency distribution)")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Build competitor analysis workbook")
    parser.add_argument("--msa-activity", required=True)
    parser.add_argument("--properties", required=True)
    parser.add_argument("--rentals", required=True)
    parser.add_argument("--for-sale", required=True)
    parser.add_argument("--transactions", required=True)
    parser.add_argument("--rpi", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--investor", default="Investor")
    build_workbook(parser.parse_args())
