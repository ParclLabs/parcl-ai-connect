#!/usr/bin/env python3
# /// script
# requires-python = ">=3.11"
# dependencies = [
#     "pandas>=2.0",
#     "openpyxl>=3.1",
# ]
# ///
"""Step 5: write the final four-tab xlsx from the profiles CSV.

Input:  build_profiles.py output (CSV with one row per PM).
Output: xlsx with four tabs:
    - PM Profiles                        : third_party category, ranked by active_rentals
    - Brokerages                         : brokerage category (separate — not silently ranked as PMs)
    - Institutional Landlords            : institutional_landlord category
    - Rental Software & Platforms        : rental_software category (Zumper,
                                           AppFolio, Zillow, etc. — visible as
                                           leasing-activity surface, not PMs)

Each tab has autofilter + frozen header row. No formulas.

CLI:
    python build_spreadsheet.py --input PROFILES.csv --output PM_LIST.xlsx
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


# Column order for each tab. Chosen for readability: name first, scale next,
# geography, then metrics, then flags/contact.
COLUMN_ORDER = [
    "pm_display_name",
    "category",
    "active_rentals",
    "num_states",
    "states",
    "num_msas",
    "top_msas",
    "num_cities",
    "num_zips",
    "avg_rent",
    "median_rent",
    "avg_sqft",
    "avg_dom",
    "median_dom",
    "pct_single_family",
    "pct_condo",
    "pct_townhouse",
    "avg_mri",
    "pct_desperate_to_lease",
    "pct_motivated",
    "pct_algo_pricing",
    "is_pm_flagged",
    "primary_company_type",
    "contact_fill_rate",
    "representative_email",
    "representative_phone",
    "top_resolution_source",
    "pm_normalized",
]


def _apply_style(ws) -> None:
    """Style the header row and set reasonable column widths."""
    if ws.max_row == 0:
        return
    # Header styling.
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
    # Column widths: simple heuristic based on header length.
    for col in range(1, ws.max_column + 1):
        header = str(ws.cell(row=1, column=col).value or "")
        width = max(12, min(40, len(header) + 2))
        # Give more room to name / states / top_msas / email columns.
        if header in ("pm_display_name", "states", "top_msas"):
            width = 42
        elif header in ("representative_email", "representative_phone"):
            width = 28
        ws.column_dimensions[get_column_letter(col)].width = width
    # Freeze header, add autofilter.
    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"


def _write_tab(wb: Workbook, title: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(title=title[:31])  # xlsx sheet name max 31 chars

    # Preserve column order; drop columns not in the order for cleanliness.
    cols = [c for c in COLUMN_ORDER if c in df.columns]
    df = df[cols]

    # Write header.
    ws.append(list(df.columns))
    # Write rows.
    for row in df.itertuples(index=False, name=None):
        # openpyxl handles pandas NaN as blank; convert to avoid "nan" strings.
        cleaned = [None if pd.isna(v) else v for v in row]
        ws.append(cleaned)

    _apply_style(ws)


def _write_summary_tab(wb: Workbook, df: pd.DataFrame) -> None:
    """Summary tab: one row per category showing unique PMs + unique
    listings represented on that tab."""
    ws = wb.create_sheet(title="Summary", index=0)

    def _stats(cat: str) -> tuple[int, int]:
        sub = df[df["category"] == cat] if "category" in df.columns else df.iloc[0:0]
        pms = len(sub)
        listings = int(sub["active_rentals"].sum()) if "active_rentals" in sub.columns and pms else 0
        return pms, listings

    tp_pms, tp_list = _stats("third_party")
    br_pms, br_list = _stats("brokerage")
    in_pms, in_list = _stats("institutional_landlord")
    sw_pms, sw_list = _stats("rental_software")

    rows = [
        ["Property Manager Intel — Summary"],
        [],
        ["Tab", "Unique PMs", "Unique Listings"],
        ["PM Profiles (third-party)", tp_pms, tp_list],
        ["Brokerages", br_pms, br_list],
        ["Institutional Landlords", in_pms, in_list],
        ["Rental Software & Platforms", sw_pms, sw_list],
    ]
    for r in rows:
        ws.append(r)

    # Styling.
    ws["A1"].font = Font(bold=True, size=14)
    # Header row for the category table.
    for col in range(1, 4):
        c = ws.cell(row=3, column=col)
        c.font = Font(bold=True)
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 18


def build_workbook(profiles: pd.DataFrame) -> Workbook:
    wb = Workbook()
    # Remove the default sheet; we add our own.
    default = wb.active
    wb.remove(default)

    third = profiles[profiles["category"] == "third_party"]
    brokerages = profiles[profiles["category"] == "brokerage"]
    institutional = profiles[profiles["category"] == "institutional_landlord"]
    software = profiles[profiles["category"] == "rental_software"]

    _write_summary_tab(wb, profiles)
    _write_tab(wb, "PM Profiles", third)
    _write_tab(wb, "Brokerages", brokerages)
    _write_tab(wb, "Institutional Landlords", institutional)
    # xlsx sheet name max 31 chars. Keep the sheet tab and the Summary
    # row name in sync so users don't see two labels for the same thing.
    _write_tab(wb, "Rental Software & Platforms", software)

    return wb


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--input", required=True, help="profiles CSV from build_profiles")
    ap.add_argument("--output", required=True, help="output xlsx")
    args = ap.parse_args()

    profiles = pd.read_csv(args.input, low_memory=False)
    wb = build_workbook(profiles)
    wb.save(args.output)

    total = len(profiles)
    out_path = Path(args.output).resolve()
    print(f"build_spreadsheet: pms={total} file={out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
