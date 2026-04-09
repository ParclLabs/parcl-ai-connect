#!/usr/bin/env python3
"""
build_workbook.py — Broker Analytics Motivated Seller Workbook Generator

Reads a CSV of motivated seller properties (from Parcl Labs), filters to luxury
listings (>$2M), and produces a 5-sheet Excel workbook analyzing agent/brokerage
performance ranked by motivated seller distress signals.

Usage:
    python3 build_workbook.py <input_csv> <output_xlsx> --market "Market Name"
"""

import argparse
import sys
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, numbers
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Color palette
# ---------------------------------------------------------------------------
NAVY_BG = PatternFill("solid", fgColor="1F4E79")
WHITE_BOLD_14 = Font(name="Arial", size=14, bold=True, color="FFFFFF")
WHITE_BOLD_10 = Font(name="Arial", size=10, bold=True, color="FFFFFF")
GRAY_10 = Font(name="Arial", size=10, color="666666")
RED_BOLD_10 = Font(name="Arial", size=10, bold=True, color="C00000")
NAVY_BOLD_10 = Font(name="Arial", size=10, bold=True, color="1F4E79")
BOLD_10 = Font(name="Arial", size=10, bold=True)
NORMAL_10 = Font(name="Arial", size=10)
LIGHT_BLUE_BG = PatternFill("solid", fgColor="D6E4F0")
BLUE_HEADER_BG = PatternFill("solid", fgColor="2E75B6")
ORANGE_HEADER_BG = PatternFill("solid", fgColor="ED7D31")
DARKRED_HEADER_BG = PatternFill("solid", fgColor="C00000")
FIRE_SALE_BG = PatternFill("solid", fgColor="FFC7CE")
MOTIVATED_BG = PatternFill("solid", fgColor="FFEB9C")
CREAM_BG = PatternFill("solid", fgColor="FFF2CC")
VS_MARKET_POS = {"font": Font(name="Arial", size=10, color="9C0006"), "fill": PatternFill("solid", fgColor="FFC7CE")}
VS_MARKET_NEG = {"font": Font(name="Arial", size=10, color="006100"), "fill": PatternFill("solid", fgColor="C6EFCE")}

FMT_CURRENCY = '$#,##0'
FMT_PCT = '0.0%'
FMT_SCORE = '0.00'
FMT_INT = '0'
FMT_FLOAT1 = '0.0'
FMT_VS_MARKET = '+0.00;-0.00;0.00'
WRAP = Alignment(wrap_text=True, vertical="top")


def safe(val):
    """Convert NaN / None to empty string for Excel."""
    if pd.isna(val):
        return ""
    return val


def set_col_widths(ws, widths: dict):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


def merge_title(ws, row, last_col, text, font=WHITE_BOLD_14, fill=NAVY_BG):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = font
    cell.fill = fill
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(2, last_col + 1):
        ws.cell(row=row, column=c).fill = fill


def write_header_row(ws, row, headers, fill, font=WHITE_BOLD_10):
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def apply_row_fill(ws, row, ncols, fill):
    for c in range(1, ncols + 1):
        ws.cell(row=row, column=c).fill = fill


def fmt_cell(ws, row, col, fmt):
    ws.cell(row=row, column=col).number_format = fmt


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_executive_summary(wb, df, market, avg_score):
    ws = wb.active
    ws.title = "Executive Summary"

    today_str = date.today().strftime("%Y-%m-%d")

    # Title
    merge_title(ws, 1, 8, f"{market.upper()} FOR-SALE MARKET: MOTIVATED SELLER ANALYSIS BY BROKER/AGENT")

    # Subtitle
    ws.merge_cells("A2:H2")
    c = ws.cell(row=2, column=1, value=f"Properties Listed Over $2M | Data as of {today_str} | Source: Parcl Labs")
    c.font = GRAY_10
    for col in range(2, 9):
        ws.cell(row=2, column=col)  # ensure merge

    # Section header — Market Snapshot
    ws.merge_cells("A4:C4")
    c = ws.cell(row=4, column=1, value="MARKET SNAPSHOT")
    c.font = NAVY_BOLD_10
    c.fill = LIGHT_BLUE_BG
    for col in range(2, 4):
        ws.cell(row=4, column=col).fill = LIGHT_BLUE_BG

    labels_values = [
        ("Total Listings Over $2M", len(df)),
        ("Listings with Agent Data", int(df["agent_name"].notna().sum())),
        ("Unique Brokerages", int(df["agent_business"].nunique())),
        ("Unique Agents", int(df["agent_name"].nunique())),
        ("Median Listing Price", df["latest_listing_price"].median()),
        ("Avg Days on Market", df["days_on_market"].mean()),
        ("Avg Motivated Seller Score", avg_score),
        ("Fire Sale Properties", int((df["motivated_seller_index_label"] == "fire_sale").sum())),
        ("Motivated Properties", int((df["motivated_seller_index_label"] == "motivated").sum())),
        ("Stubborn Properties", int((df["motivated_seller_index_label"] == "stubborn").sum())),
        ("Neutral Properties", int((df["motivated_seller_index_label"] == "neutral").sum())),
    ]
    for i, (label, val) in enumerate(labels_values):
        r = 5 + i
        ws.cell(row=r, column=1, value=label).font = BOLD_10
        vc = ws.cell(row=r, column=2, value=val)
        vc.font = NORMAL_10
        if "Price" in label:
            vc.number_format = FMT_CURRENCY
        elif "Score" in label:
            vc.number_format = FMT_SCORE
        elif "Days" in label:
            vc.number_format = FMT_INT

    # Label distribution section
    ws.merge_cells("A17:E17")
    c = ws.cell(row=17, column=1, value="MOTIVATED SELLER LABEL DISTRIBUTION")
    c.font = NAVY_BOLD_10
    c.fill = LIGHT_BLUE_BG
    for col in range(2, 6):
        ws.cell(row=17, column=col).fill = LIGHT_BLUE_BG

    dist_headers = ["Label", "Count", "% of Total", "Avg Price", "Avg DOM"]
    for i, h in enumerate(dist_headers, 1):
        ws.cell(row=18, column=i, value=h).font = BOLD_10

    total = len(df)
    for i, label in enumerate(["fire_sale", "motivated", "stubborn", "neutral"]):
        r = 19 + i
        subset = df[df["motivated_seller_index_label"] == label]
        cnt = len(subset)
        ws.cell(row=r, column=1, value=label).font = NORMAL_10
        ws.cell(row=r, column=2, value=cnt).font = NORMAL_10
        pct_cell = ws.cell(row=r, column=3, value=cnt / total if total else 0)
        pct_cell.number_format = FMT_PCT
        pct_cell.font = NORMAL_10
        price_cell = ws.cell(row=r, column=4, value=subset["latest_listing_price"].mean() if cnt else 0)
        price_cell.number_format = FMT_CURRENCY
        price_cell.font = NORMAL_10
        dom_cell = ws.cell(row=r, column=5, value=subset["days_on_market"].mean() if cnt else 0)
        dom_cell.number_format = FMT_INT
        dom_cell.font = NORMAL_10

    # Key Insight
    ws.merge_cells("A24:H24")
    c = ws.cell(row=24, column=1, value="KEY INSIGHT")
    c.font = NAVY_BOLD_10
    c.fill = LIGHT_BLUE_BG
    for col in range(2, 9):
        ws.cell(row=24, column=col).fill = LIGHT_BLUE_BG

    insight = (
        f"Market avg motivated seller score is {avg_score:.2f}. "
        "Brokerages/agents scoring ABOVE this average have listings that are sitting longer, "
        "cutting prices more, and signaling more seller distress — a potential indicator of "
        "overpricing or underperformance relative to the market."
    )
    ws.merge_cells("A25:H26")
    c = ws.cell(row=25, column=1, value=insight)
    c.font = NORMAL_10
    c.alignment = WRAP

    set_col_widths(ws, {"A": 32, "B": 22, "C": 16, "D": 13, "E": 13, "F": 13, "G": 13, "H": 13})


def build_all_properties(wb, df, market):
    ws = wb.create_sheet("All Properties Over $2M")

    headers = [
        "Parcl Property ID", "Address", "City", "ZIP", "Type", "Beds", "Baths",
        "Sq Ft", "Year Built", "Current Price", "Initial Price", "Days on Market",
        "Price Cuts", "% Price Change", "Motivated Score", "Motivated Label",
        "Agent Name", "Brokerage", "Agent Email", "Agent Phone",
        "Owner Entity", "Original Purchase Price", "Unrealized P&L",
    ]
    fields = [
        "parcl_property_id", "address1", "city", "zip5", "property_type",
        "bedrooms", "bathrooms", "sq_ft", "year_built",
        "latest_listing_price", "initial_listing_price", "days_on_market",
        "num_price_cuts", "percent_change_listing_price",
        "motivated_seller_index_value", "motivated_seller_index_label",
        "agent_name", "agent_business", "agent_email", "agent_phone",
        "entity_name", "original_purchase_price", "unrealized_net",
    ]
    currency_cols = {10, 11, 22, 23}
    pct_cols = {14}
    score_cols = {15}
    ncols = len(headers)

    # Sort
    df_sorted = df.sort_values(
        ["motivated_seller_index_value", "days_on_market"],
        ascending=[False, False],
    ).reset_index(drop=True)

    merge_title(ws, 1, ncols, f"ALL {market.upper()} PROPERTIES FOR SALE OVER $2M")
    write_header_row(ws, 2, headers, BLUE_HEADER_BG)

    for idx, row in df_sorted.iterrows():
        r = idx + 3
        for ci, field in enumerate(fields, 1):
            val = safe(row.get(field, ""))
            # Convert percent_change_listing_price to decimal for Excel pct format
            if ci in pct_cols and val != "" and val is not None:
                try:
                    val = float(val) / 100.0
                except (ValueError, TypeError):
                    pass
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font = NORMAL_10
            if ci in currency_cols:
                cell.number_format = FMT_CURRENCY
            elif ci in pct_cols:
                cell.number_format = FMT_PCT
            elif ci in score_cols:
                cell.number_format = FMT_SCORE

        label = safe(row.get("motivated_seller_index_label", ""))
        if label == "fire_sale":
            apply_row_fill(ws, r, ncols, FIRE_SALE_BG)
        elif label == "motivated":
            apply_row_fill(ws, r, ncols, MOTIVATED_BG)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(ncols)}{len(df_sorted) + 2}"

    widths = {
        "A": 18, "B": 35, "C": 16, "D": 8, "E": 14, "F": 6, "G": 13,
        "H": 10, "I": 13, "J": 15, "K": 13, "L": 13, "M": 10, "N": 12,
        "O": 13, "P": 13, "Q": 25, "R": 40, "S": 28, "T": 16, "U": 25,
        "V": 18, "W": 16,
    }
    set_col_widths(ws, widths)


def build_brokerage_analysis(wb, df, market, avg_score):
    ws = wb.create_sheet("Brokerage Analysis")

    ncols = 12
    merge_title(ws, 1, ncols, "BROKERAGE PERFORMANCE: RANKED BY MOTIVATED SELLER SCORE (UNDERPERFORMERS AT TOP)")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(row=2, column=1, value=f"Market Avg Motivated Score: {avg_score:.2f} | Brokerages above this line are underperforming the market")
    c.font = RED_BOLD_10
    for col in range(2, ncols + 1):
        ws.cell(row=2, column=col)

    headers = [
        "Brokerage", "# Listings", "Avg Motivated Score", "vs Market Avg",
        "Avg Days on Market", "Avg Price Cuts", "Avg % Price Change",
        "Total Listing Value", "Avg Listing Price", "Fire Sale", "Motivated",
        "Neutral + Stubborn",
    ]
    write_header_row(ws, 3, headers, ORANGE_HEADER_BG)

    # Filter to rows with brokerage data
    bdf = df[df["agent_business"].notna() & (df["agent_business"] != "")].copy()
    agg = bdf.groupby("agent_business").agg(
        count=("latest_listing_price", "size"),
        avg_score=("motivated_seller_index_value", "mean"),
        avg_dom=("days_on_market", "mean"),
        avg_cuts=("num_price_cuts", "mean"),
        avg_pct=("percent_change_listing_price", "mean"),
        total_val=("latest_listing_price", "sum"),
        avg_price=("latest_listing_price", "mean"),
        fire=("motivated_seller_index_label", lambda x: (x == "fire_sale").sum()),
        motiv=("motivated_seller_index_label", lambda x: (x == "motivated").sum()),
        neutral_stub=("motivated_seller_index_label", lambda x: x.isin(["neutral", "stubborn"]).sum()),
    ).reset_index().sort_values("avg_score", ascending=False)

    for idx, row in agg.iterrows():
        r = idx + 4 if idx == agg.index[0] else None
    # Re-index for clean row writing
    for i, (_, row) in enumerate(agg.iterrows()):
        r = 4 + i
        delta = row["avg_score"] - avg_score

        ws.cell(row=r, column=1, value=row["agent_business"]).font = BOLD_10
        ws.cell(row=r, column=2, value=int(row["count"])).font = NORMAL_10
        ws.cell(row=r, column=3, value=row["avg_score"]).font = NORMAL_10
        ws.cell(row=r, column=3).number_format = FMT_SCORE
        dc = ws.cell(row=r, column=4, value=delta)
        dc.number_format = FMT_VS_MARKET
        ws.cell(row=r, column=5, value=row["avg_dom"]).number_format = FMT_INT
        ws.cell(row=r, column=5).font = NORMAL_10
        ws.cell(row=r, column=6, value=row["avg_cuts"]).number_format = FMT_FLOAT1
        ws.cell(row=r, column=6).font = NORMAL_10
        pct_cell = ws.cell(row=r, column=7, value=row["avg_pct"] / 100.0 if row["avg_pct"] else 0)
        pct_cell.number_format = FMT_PCT
        pct_cell.font = NORMAL_10
        ws.cell(row=r, column=8, value=row["total_val"]).number_format = FMT_CURRENCY
        ws.cell(row=r, column=8).font = NORMAL_10
        ws.cell(row=r, column=9, value=row["avg_price"]).number_format = FMT_CURRENCY
        ws.cell(row=r, column=9).font = NORMAL_10
        ws.cell(row=r, column=10, value=int(row["fire"])).font = NORMAL_10
        ws.cell(row=r, column=11, value=int(row["motiv"])).font = NORMAL_10
        ws.cell(row=r, column=12, value=int(row["neutral_stub"])).font = NORMAL_10

        if row["avg_score"] > avg_score:
            apply_row_fill(ws, r, ncols, CREAM_BG)

        if delta > 0:
            dc.font = VS_MARKET_POS["font"]
            dc.fill = VS_MARKET_POS["fill"]
        else:
            dc.font = VS_MARKET_NEG["font"]
            dc.fill = VS_MARKET_NEG["fill"]

    ws.freeze_panes = "A4"
    last_row = 3 + len(agg)
    ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}{last_row}"

    set_col_widths(ws, {
        "A": 48, "B": 10, "C": 16, "D": 14, "E": 16, "F": 12,
        "G": 14, "H": 18, "I": 16, "J": 10, "K": 13, "L": 16,
    })


def build_agent_analysis(wb, df, market, avg_score):
    ws = wb.create_sheet("Agent Analysis")

    ncols = 11
    merge_title(ws, 1, ncols, "AGENT PERFORMANCE: RANKED BY MOTIVATED SELLER SCORE (UNDERPERFORMERS AT TOP)")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(row=2, column=1, value=f"Market Avg Motivated Score: {avg_score:.2f} | Agents above this line have more distressed listings")
    c.font = RED_BOLD_10
    for col in range(2, ncols + 1):
        ws.cell(row=2, column=col)

    headers = [
        "Agent", "Brokerage", "# Listings", "Avg Motivated Score",
        "vs Market Avg", "Avg Days on Market", "Avg Price Cuts",
        "Avg % Price Change", "Total Value", "Fire Sales", "Motivated",
    ]
    write_header_row(ws, 3, headers, DARKRED_HEADER_BG)

    adf = df[df["agent_name"].notna() & (df["agent_name"] != "")].copy()
    agg = adf.groupby(["agent_name", "agent_business"]).agg(
        count=("latest_listing_price", "size"),
        avg_score=("motivated_seller_index_value", "mean"),
        avg_dom=("days_on_market", "mean"),
        avg_cuts=("num_price_cuts", "mean"),
        avg_pct=("percent_change_listing_price", "mean"),
        total_val=("latest_listing_price", "sum"),
        fire=("motivated_seller_index_label", lambda x: (x == "fire_sale").sum()),
        motiv=("motivated_seller_index_label", lambda x: (x == "motivated").sum()),
    ).reset_index().sort_values("avg_score", ascending=False)

    freeze_row = None
    for i, (_, row) in enumerate(agg.iterrows()):
        r = 4 + i
        delta = row["avg_score"] - avg_score

        ws.cell(row=r, column=1, value=row["agent_name"]).font = BOLD_10
        ws.cell(row=r, column=2, value=safe(row["agent_business"])).font = NORMAL_10
        ws.cell(row=r, column=3, value=int(row["count"])).font = NORMAL_10
        ws.cell(row=r, column=4, value=row["avg_score"]).number_format = FMT_SCORE
        ws.cell(row=r, column=4).font = NORMAL_10
        dc = ws.cell(row=r, column=5, value=delta)
        dc.number_format = FMT_VS_MARKET
        ws.cell(row=r, column=6, value=row["avg_dom"]).number_format = FMT_INT
        ws.cell(row=r, column=6).font = NORMAL_10
        ws.cell(row=r, column=7, value=row["avg_cuts"]).number_format = FMT_FLOAT1
        ws.cell(row=r, column=7).font = NORMAL_10
        pct_cell = ws.cell(row=r, column=8, value=row["avg_pct"] / 100.0 if row["avg_pct"] else 0)
        pct_cell.number_format = FMT_PCT
        pct_cell.font = NORMAL_10
        ws.cell(row=r, column=9, value=row["total_val"]).number_format = FMT_CURRENCY
        ws.cell(row=r, column=9).font = NORMAL_10
        ws.cell(row=r, column=10, value=int(row["fire"])).font = NORMAL_10
        ws.cell(row=r, column=11, value=int(row["motiv"])).font = NORMAL_10

        if row["avg_score"] > avg_score:
            apply_row_fill(ws, r, ncols, CREAM_BG)
        elif freeze_row is None:
            freeze_row = r + 2  # a couple rows below the transition

        if delta > 0:
            dc.font = VS_MARKET_POS["font"]
            dc.fill = VS_MARKET_POS["fill"]
        else:
            dc.font = VS_MARKET_NEG["font"]
            dc.fill = VS_MARKET_NEG["fill"]

    if freeze_row and freeze_row <= 3 + len(agg):
        ws.freeze_panes = f"A{freeze_row}"
    else:
        ws.freeze_panes = "A4"

    last_row = 3 + len(agg)
    ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}{last_row}"

    set_col_widths(ws, {
        "A": 20, "B": 48, "C": 10, "D": 16, "E": 14,
        "F": 16, "G": 12, "H": 14, "I": 16, "J": 10, "K": 13,
    })


def build_fire_sale_detail(wb, df, market):
    ws = wb.create_sheet("Fire Sale & Motivated")

    distressed = df[df["motivated_seller_index_label"].isin(["fire_sale", "motivated"])].copy()
    distressed = distressed.sort_values(
        ["motivated_seller_index_value", "days_on_market"],
        ascending=[False, False],
    ).reset_index(drop=True)

    ncols = 15
    merge_title(ws, 1, ncols, f"FIRE SALE & MOTIVATED PROPERTIES — {market.upper()}", fill=DARKRED_HEADER_BG)

    headers = [
        "Parcl Property ID", "Address", "City", "ZIP", "Current Price",
        "Initial Price", "Days on Market", "Price Cuts", "% Price Change",
        "Motivated Score", "Label", "Agent", "Brokerage", "Owner Entity",
        "Unrealized P&L",
    ]
    fields = [
        "parcl_property_id", "address1", "city", "zip5",
        "latest_listing_price", "initial_listing_price", "days_on_market",
        "num_price_cuts", "percent_change_listing_price",
        "motivated_seller_index_value", "motivated_seller_index_label",
        "agent_name", "agent_business", "entity_name", "unrealized_net",
    ]
    currency_cols = {5, 6, 15}
    pct_cols = {9}
    score_cols = {10}

    write_header_row(ws, 2, headers, DARKRED_HEADER_BG)

    for idx, row in distressed.iterrows():
        r = idx + 3
        for ci, field in enumerate(fields, 1):
            val = safe(row.get(field, ""))
            if ci in pct_cols and val != "" and val is not None:
                try:
                    val = float(val) / 100.0
                except (ValueError, TypeError):
                    pass
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font = NORMAL_10
            if ci in currency_cols:
                cell.number_format = FMT_CURRENCY
            elif ci in pct_cols:
                cell.number_format = FMT_PCT
            elif ci in score_cols:
                cell.number_format = FMT_SCORE

        label = safe(row.get("motivated_seller_index_label", ""))
        if label == "fire_sale":
            apply_row_fill(ws, r, ncols, FIRE_SALE_BG)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(ncols)}{len(distressed) + 2}"

    set_col_widths(ws, {
        "A": 18, "B": 35, "C": 16, "D": 8, "E": 15, "F": 13,
        "G": 13, "H": 10, "I": 12, "J": 13, "K": 12, "L": 25,
        "M": 42, "N": 25, "O": 16,
    })


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Build Broker Analytics workbook")
    parser.add_argument("input_csv", help="Path to motivated seller CSV")
    parser.add_argument("output_xlsx", help="Path for output Excel file")
    parser.add_argument("--market", required=True, help="Market display name")
    parser.add_argument("--min-price", type=float, default=2_000_000, help="Minimum listing price filter (default: 2000000)")
    args = parser.parse_args()

    csv_path = Path(args.input_csv)
    if not csv_path.exists():
        print(f"ERROR: Input CSV not found: {csv_path}", file=sys.stderr)
        sys.exit(1)

    print(f"Reading {csv_path}...")
    df = pd.read_csv(csv_path)
    print(f"  Total rows: {len(df)}")

    # Ensure numeric columns
    for col in ["latest_listing_price", "initial_listing_price", "days_on_market",
                 "num_price_cuts", "percent_change_listing_price",
                 "motivated_seller_index_value", "bedrooms", "bathrooms",
                 "sq_ft", "year_built", "original_purchase_price", "unrealized_net"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Filter luxury
    df = df[df["latest_listing_price"] > args.min_price].reset_index(drop=True)
    print(f"  After filtering >${args.min_price:,.0f}: {len(df)} properties")

    if len(df) == 0:
        print("WARNING: No properties above price threshold. Generating empty workbook.", file=sys.stderr)

    avg_score = df["motivated_seller_index_value"].mean() if len(df) > 0 else 0

    wb = Workbook()
    build_executive_summary(wb, df, args.market, avg_score)
    build_all_properties(wb, df, args.market)
    build_brokerage_analysis(wb, df, args.market, avg_score)
    build_agent_analysis(wb, df, args.market, avg_score)
    build_fire_sale_detail(wb, df, args.market)

    out_path = Path(args.output_xlsx)
    wb.save(str(out_path))
    print(f"Workbook saved to {out_path}")
    print(f"  Sheets: {wb.sheetnames}")
    print(f"  Properties: {len(df)}")
    print(f"  Market avg score: {avg_score:.2f}")


if __name__ == "__main__":
    main()
